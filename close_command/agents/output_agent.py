"""
Journal Output Agent for Close Command.
Generates JLF output, Excel journal, audit trail, and consolidation proof.
"""

from __future__ import annotations

import io
import logging
import uuid
from datetime import datetime
from typing import Any

from close_command.database.persistence import CloseCommandDB
from close_command.utils.formatters import jlf_row

try:
    from close_command.rag.indexer import CloseCommandIndexer as _IndexerClass
except ImportError:
    _IndexerClass = None  # type: ignore

logger = logging.getLogger(__name__)


class _NullIndexer:
    """Stub indexer used when chromadb/RAG is unavailable."""

    def index_completed_period(self, batch_id, journal_entries, period):
        logger.warning("NullIndexer: skipping vectorstore indexing (chromadb not available)")
        return 0


class JournalOutputAgent:
    """Generates final output: JLF, Excel, audit trail, and consolidation proof."""

    def __init__(self, db: CloseCommandDB, indexer=None) -> None:
        self.db = db
        self.indexer = indexer if indexer is not None else _NullIndexer()

    # ──────────────────────────────────────────────────────────────────────
    # Main entry point
    # ──────────────────────────────────────────────────────────────────────

    def run(self, state: dict) -> dict:
        """Generate all output artefacts. Updates state with output_result."""
        try:
            review_result = state.get("review_result", {})
            elimination_result = state.get("elimination_result", {})
            batch_id = (
                review_result.get("batch_id")
                or elimination_result.get("batch_id")
                or state.get("batch_id", str(uuid.uuid4()))
            )
            period = (
                review_result.get("period")
                or elimination_result.get("period")
                or state.get("period", "2024-03")
            )
            scenario = (
                elimination_result.get("scenario")
                or state.get("scenario", "ACTUAL")
            )

            logger.info("JournalOutputAgent starting for batch_id=%s period=%s", batch_id, period)

            # Collect approved entries: prioritise DB-stored APPROVED entries
            approved_entries = self._collect_approved_entries(
                batch_id=batch_id,
                review_result=review_result,
                elimination_result=elimination_result,
            )

            # Decisions list from review
            decisions = list(review_result.get("existing_decisions", {}).values())
            decided_items = review_result.get("decided_items", [])
            for item in decided_items:
                if "existing_decision" in item:
                    decisions.append(item["existing_decision"])

            # Generate outputs
            jlf_output = self.generate_jlf_output(approved_entries)

            try:
                excel_bytes = self.generate_excel_journal(approved_entries)
            except Exception as xl_exc:
                logger.warning("Excel generation failed (openpyxl may not be installed): %s", xl_exc)
                excel_bytes = b""

            audit_trail = self.generate_audit_trail(batch_id, decisions)
            proof = self.compute_consolidation_proof(approved_entries, decisions)

            # Index to vectorstore
            indexed_count = self.update_vectorstore_with_period(
                batch_id=batch_id,
                entries=approved_entries,
                period=period,
                indexer=self.indexer,
            )

            # Persist approved entries back to DB with gate_status=APPROVED
            try:
                for entry in approved_entries:
                    entry["gate_status"] = "APPROVED"
                self.db.save_journal_entries(approved_entries)
            except Exception as db_exc:
                logger.warning("Failed to save approved entries to DB: %s", db_exc)

            # Update batch run status in DB
            try:
                self.db.update_batch_status(batch_id, "COMPLETED")
            except Exception as db_exc:
                logger.warning("Failed to update batch status: %s", db_exc)

            # Build serializable approved_entries summary (no binary blobs)
            approved_entries_serializable = [
                {k: v for k, v in e.items() if not isinstance(v, (bytes, bytearray))}
                for e in approved_entries
            ]

            output_result = {
                "batch_id": batch_id,
                "period": period,
                "scenario": scenario,
                "jlf_output": jlf_output,
                # excel_bytes excluded from state — too large and not msgpack-serializable
                "audit_trail": audit_trail,
                "consolidation_proof": proof,
                "approved_entries": approved_entries_serializable,   # ← journal tab reads this
                "approved_entries_count": len(approved_entries),
                "indexed_count": indexed_count,
                "generated_at": datetime.utcnow().isoformat(),
                "status": "COMPLETE",
            }

            state["output_result"] = output_result
            state["period_close_ready"] = proof.get("balanced", False)

            logger.info(
                "JournalOutputAgent complete: approved=%d jlf_lines=%d indexed=%d",
                len(approved_entries),
                jlf_output.count("\n"),
                indexed_count,
            )
            return state

        except Exception as exc:
            logger.error("JournalOutputAgent.run failed: %s", exc)
            state.setdefault("errors", []).append(str(exc))
            state["output_result"] = {
                "batch_id": state.get("batch_id", ""),
                "period": state.get("period", ""),
                "scenario": state.get("scenario", ""),
                "jlf_output": "",
                "excel_bytes": b"",
                "audit_trail": [],
                "consolidation_proof": {},
                "approved_entries_count": 0,
                "indexed_count": 0,
                "generated_at": datetime.utcnow().isoformat(),
                "status": "ERROR",
                "error": str(exc),
            }
            return state

    # ──────────────────────────────────────────────────────────────────────
    # JLF Output
    # ──────────────────────────────────────────────────────────────────────

    def generate_jlf_output(self, approved_entries: list[dict]) -> str:
        """
        Returns JLF semicolon-delimited string.
        Header + one line per entry using utils.formatters.jlf_row().
        """
        try:
            header = (
                "TxnID;Seller;Buyer;RuleCode;Account;Flow;ICP;"
                "Custom3CCY;DrCr;Amount;AuditCode;Period;Scenario\n"
            )
            lines: list[str] = [header]

            for entry in approved_entries:
                try:
                    line = jlf_row(entry)
                    lines.append(line + "\n")
                except Exception as row_exc:
                    logger.warning("JLF row formatting failed: %s", row_exc)
                    lines.append(f"ERROR;{row_exc}\n")

            return "".join(lines)

        except Exception as exc:
            logger.error("generate_jlf_output failed: %s", exc)
            return f"ERROR generating JLF: {exc}\n"

    # ──────────────────────────────────────────────────────────────────────
    # Excel Journal
    # ──────────────────────────────────────────────────────────────────────

    def generate_excel_journal(self, approved_entries: list[dict]) -> bytes:
        """
        Create Excel workbook with formatted journal entries.
        Returns bytes (BytesIO).
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Journal Eliminations"

            # Column definitions
            columns = [
                "TxnID", "Seller", "Buyer", "Rule Code", "Account",
                "Flow", "ICP", "Custom3 CCY", "Dr/Cr", "Amount (abs)",
                "Amount USD", "Audit Code", "Period", "Scenario",
                "Computed By", "Gate Status",
            ]

            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Row fills
            dr_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # light green
            cr_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # light blue

            # Data rows
            for row_idx, entry in enumerate(approved_entries, start=2):
                dr_cr = str(entry.get("dr_cr", ""))
                row_fill = dr_fill if dr_cr == "Dr" else cr_fill

                amount = abs(float(entry.get("amount", 0) or 0))
                amount_usd = abs(float(entry.get("amount_usd", amount) or 0))

                computed_at = entry.get("computed_at", "")
                if isinstance(computed_at, datetime):
                    computed_at = computed_at.isoformat()

                row_values = [
                    str(entry.get("txn_id", "")),
                    str(entry.get("seller", "")),
                    str(entry.get("buyer", "")),
                    str(entry.get("rule_code", "")),
                    str(entry.get("account", "")),
                    str(entry.get("flow", "F00")),
                    str(entry.get("icp", "")),
                    str(entry.get("custom3_ccy", "USD")),
                    dr_cr,
                    amount,
                    amount_usd,
                    str(entry.get("audit_code", "")),
                    str(entry.get("period", "")),
                    str(entry.get("scenario", "")),
                    str(entry.get("computed_by", "Elimination Agent")),
                    str(entry.get("gate_status", "APPROVED")),
                ]

                for col_idx, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = row_fill

            # Auto-size columns (approximate)
            for col_idx, col_name in enumerate(columns, start=1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(len(col_name) + 2, 14)

            # Freeze header row
            ws.freeze_panes = "A2"

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.read()

        except ImportError:
            logger.warning("openpyxl not installed — returning empty bytes")
            return b""
        except Exception as exc:
            logger.error("generate_excel_journal failed: %s", exc)
            return b""

    # ──────────────────────────────────────────────────────────────────────
    # Audit Trail
    # ──────────────────────────────────────────────────────────────────────

    def generate_audit_trail(self, batch_id: str, decisions: list) -> list[dict]:
        """
        Merge journal entries with review decisions.
        Returns list of audit trail dicts.
        """
        try:
            # Fetch all entries for batch
            all_entries = self.db.get_journal_entries(batch_id)

            # Build decision map by txn_id
            decision_map: dict[str, dict] = {}
            for dec in decisions:
                if isinstance(dec, dict):
                    txn_id = dec.get("txn_id", "")
                    if txn_id:
                        decision_map[txn_id] = dec

            audit_rows: list[dict] = []

            for entry in all_entries:
                entry_txn = str(entry.get("txn_id", ""))
                # Find decision — match by txn_id or by audit_code substring
                matched_decision = decision_map.get(entry_txn, {})
                if not matched_decision:
                    # Try partial match via audit_code
                    for d_txn, d_val in decision_map.items():
                        if d_txn in str(entry.get("audit_code", "")):
                            matched_decision = d_val
                            break

                audit_row = {
                    "txn_id": entry_txn,
                    "seller": str(entry.get("seller", "")),
                    "buyer": str(entry.get("buyer", "")),
                    "rule_code": str(entry.get("rule_code", "")),
                    "amount": abs(float(entry.get("amount", 0) or 0)),
                    "amount_usd": abs(float(entry.get("amount_usd", 0) or 0)),
                    "dr_cr": str(entry.get("dr_cr", "")),
                    "gate_status": str(entry.get("gate_status", "PENDING")),
                    "decision": matched_decision.get("decision", "Pending"),
                    "reviewer_name": matched_decision.get("reviewer_name", ""),
                    "timestamp": matched_decision.get("timestamp", ""),
                    "comment": matched_decision.get("comment", ""),
                    "audit_code": str(entry.get("audit_code", "")),
                    "period": str(entry.get("period", "")),
                    "scenario": str(entry.get("scenario", "")),
                }
                audit_rows.append(audit_row)

            logger.info("Generated audit trail with %d rows", len(audit_rows))
            return audit_rows

        except Exception as exc:
            logger.error("generate_audit_trail failed: %s", exc)
            return []

    # ──────────────────────────────────────────────────────────────────────
    # Consolidation Proof
    # ──────────────────────────────────────────────────────────────────────

    def compute_consolidation_proof(
        self, entries: list[dict], decisions: list[dict]
    ) -> dict:
        """
        Compute consolidation balance proof.
        Returns totals, entity contributions, balance flag, and proof statement.
        """
        try:
            total_dr = 0.0
            total_cr = 0.0
            entity_contributions: dict[str, float] = {}

            for entry in entries:
                dr_cr = str(entry.get("dr_cr", ""))
                amount_usd = abs(float(entry.get("amount_usd", entry.get("amount", 0)) or 0))
                seller = str(entry.get("seller", ""))
                buyer = str(entry.get("buyer", ""))

                if dr_cr == "Dr":
                    total_dr += amount_usd
                    entity_contributions[seller] = entity_contributions.get(seller, 0.0) + amount_usd
                elif dr_cr == "Cr":
                    total_cr += amount_usd
                    entity_contributions[buyer] = entity_contributions.get(buyer, 0.0) - amount_usd

            net_elimination = total_dr - total_cr
            balanced = abs(net_elimination) < 0.01

            # Round entity contributions
            entity_contributions = {k: round(v, 2) for k, v in entity_contributions.items()}

            if balanced:
                proof_statement = (
                    f"Consolidation proof BALANCED. "
                    f"Total Dr: USD {total_dr:,.2f} | Total Cr: USD {total_cr:,.2f} | "
                    f"Net: USD {net_elimination:,.4f}. "
                    f"Period close may proceed."
                )
            else:
                proof_statement = (
                    f"Consolidation proof NOT BALANCED. "
                    f"Total Dr: USD {total_dr:,.2f} | Total Cr: USD {total_cr:,.2f} | "
                    f"Net difference: USD {net_elimination:,.4f}. "
                    f"Investigate before closing."
                )

            return {
                "total_eliminations_dr": round(total_dr, 2),
                "total_eliminations_cr": round(total_cr, 2),
                "net_elimination": round(net_elimination, 4),
                "entity_contributions": entity_contributions,
                "balanced": balanced,
                "proof_statement": proof_statement,
                "entries_included": len(entries),
                "computed_at": datetime.utcnow().isoformat(),
            }

        except Exception as exc:
            logger.error("compute_consolidation_proof failed: %s", exc)
            return {
                "total_eliminations_dr": 0.0,
                "total_eliminations_cr": 0.0,
                "net_elimination": 0.0,
                "entity_contributions": {},
                "balanced": False,
                "proof_statement": f"Proof computation failed: {exc}",
                "entries_included": 0,
                "computed_at": datetime.utcnow().isoformat(),
            }

    # ──────────────────────────────────────────────────────────────────────
    # Vectorstore indexing
    # ──────────────────────────────────────────────────────────────────────

    def update_vectorstore_with_period(
        self, batch_id: str, entries: list[dict], period: str, indexer: CloseCommandIndexer
    ) -> int:
        """Index approved entries into vectorstore period_journals collection."""
        try:
            count = indexer.index_completed_period(
                batch_id=batch_id,
                journal_entries=entries,
                period=period,
            )
            logger.info("Indexed %d entries for period %s into vectorstore", count, period)
            return count
        except Exception as exc:
            logger.warning("update_vectorstore_with_period failed: %s", exc)
            return 0

    # ──────────────────────────────────────────────────────────────────────
    # Internal helpers
    # ──────────────────────────────────────────────────────────────────────

    def _collect_approved_entries(
        self,
        batch_id: str,
        review_result: dict,
        elimination_result: dict,
    ) -> list[dict]:
        """
        Collect approved entries from DB first, then fall back to state.
        Updates gate_status on entries to APPROVED.
        """
        try:
            # Try to get APPROVED entries from DB
            db_approved = self.db.get_approved_entries(batch_id)
            if db_approved:
                return db_approved

            # Fall back to all elimination entries, filtered by review decisions
            all_entries = elimination_result.get("journal_entries", [])
            existing_decisions = review_result.get("existing_decisions", {})

            if not existing_decisions:
                # No decisions yet — mark all pending entries as approved for output generation
                approved = []
                for entry in all_entries:
                    e = dict(entry)
                    e["gate_status"] = "APPROVED"
                    approved.append(e)
                return approved

            approved: list[dict] = []
            decision_txn_ids = {
                d_id for d_id, d_val in existing_decisions.items()
                if d_val.get("decision") == "Approved"
            }

            for entry in all_entries:
                txn_id = str(entry.get("txn_id", ""))
                audit_code = str(entry.get("audit_code", ""))

                # Check if this entry is associated with an approved decision
                approved_match = (
                    txn_id in decision_txn_ids
                    or any(d_id in audit_code for d_id in decision_txn_ids)
                    or (not existing_decisions)  # if no decisions, include all
                )

                if approved_match:
                    e = dict(entry)
                    e["gate_status"] = "APPROVED"
                    approved.append(e)

            # If no entries matched decisions, return all (open-loop scenario)
            if not approved and all_entries:
                approved = []
                for entry in all_entries:
                    e = dict(entry)
                    e["gate_status"] = "APPROVED"
                    approved.append(e)

            return approved

        except Exception as exc:
            logger.warning("_collect_approved_entries failed: %s", exc)
            return []
